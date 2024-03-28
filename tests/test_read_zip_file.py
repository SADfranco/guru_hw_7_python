from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv


def test_read_pdf_file():
    with ZipFile('resources/file.zip') as zip_file:
        with zip_file.open('sample-pdf-file.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            text = reader.pages[0].extract_text()
            assert 'Lorem Ipsum  is simply dummy text of the printing and typesetting industry.' in text

def test_read_csv_file():
    with ZipFile('resources/file.zip') as zip_file:
        with zip_file.open('sample-csv-file-for-testing.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            assert csvreader[1][1] == 'Canada'
            assert csvreader[2][1] == 'Germany'

def test_read_xlsx_file():
    with ZipFile('resources/file.zip') as zip_file:
        with zip_file.open('sample-xlsx-file-for-testing.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=2, column=2).value == 'Canada'
            assert sheet.cell(row=3, column=2).value == 'Germany'